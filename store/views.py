from notifications.services import (
    notify_product_added,
    notify_product_updated,
    notify_product_deleted,
    notify_store_created,
    notify_stock_adjustment,
)
from django.db.models import ProtectedError
import json
from datetime import timedelta
from decimal import Decimal
from notifications.utils import check_stock_and_notify
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.core.paginator import Paginator
from django.db import transaction
from django.db.models import Sum, Count, Q, F
from django.http import JsonResponse
from django.shortcuts import render, redirect, get_object_or_404
from django.utils import timezone
from django.views.decorators.http import require_POST

from authentication.models import CustomUser
from .forms import ProductForm, StoreForm, StockAdjustmentForm
from .models import Product, Store, Stock, StockAdjustment
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import mm
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.enums import TA_CENTER, TA_LEFT
from io import BytesIO
from datetime import datetime


@login_required
def store_home(request):
    """
    Main dashboard for the Store app.
    Shows different data depending on whether the user is Admin or store user.
    """
    user = request.user
    is_admin = getattr(user, 'can_see_all', False) or user.is_superuser

    # ---------- Base querysets with store isolation ----------
    if is_admin:
        products = Product.objects.all()
        stores = Store.objects.filter(is_active=True)
        stocks = Stock.objects.all()
    else:
        if not user.store:
            # Safety: user has no store assigned
            return render(request, 'store/dashboard.html', {
                'error': 'You are not assigned to any store.'
            })
        products = Product.objects.filter(store=user.store)
        stores = Store.objects.filter(id=user.store.id)
        stocks = Stock.objects.filter(store=user.store)

    # ---------- KPI Cards ----------
    total_products = products.filter(is_deleted=False).count()
    total_stores = stores.count()

    # Current stock value (using sell_price)
    stock_value = stocks.aggregate(
        total=Sum(F('quantity') * F('product__buy_price'))
    )['total'] or 0

    # Low stock products
    low_stock_products = []
    for product in products.filter(is_deleted=False).select_related('store'):
        stock = stocks.filter(product=product).first()
        current_qty = stock.quantity if stock else product.initial_stock
        if current_qty <= product.reorder_level:
            low_stock_products.append({
                'product': product,
                'current_stock': current_qty,
                'reorder_level': product.reorder_level
            })

    low_stock_count = len(low_stock_products)

    # Products added in last 7 days
    last_7_days = timezone.now() - timedelta(days=7)
    recent_products_count = products.filter(
        created_at__gte=last_7_days,
        is_deleted=False
    ).count() if hasattr(Product, 'created_at') else 0

    # Trash count
    if is_admin:
        trash_count = Product.all_objects.filter(is_deleted=True).count()
    else:
        trash_count = Product.all_objects.filter(
            store=user.store, is_deleted=True
        ).count()

    # Recent products (last 5)
    recent_products = products.filter(is_deleted=False).order_by('-id')[:5]

    context = {
        'is_admin': is_admin,
        'total_products': total_products,
        'total_stores': total_stores,
        'stock_value': stock_value,
        'low_stock_count': low_stock_count,
        'low_stock_products': low_stock_products[:8],  # show max 8
        'recent_products_count': recent_products_count,
        'trash_count': trash_count,
        'recent_products': recent_products,
        'current_store': user.store if not is_admin else None,
    }
    return render(request, 'store/dashboard.html', context)

@login_required
def product_list(request):
    search = request.GET.get('search')
    category = request.GET.get('category')

    products = Product.objects.for_user(request.user).select_related('store')

    if search:
        products = products.filter(product_name__icontains=search)
    if category:
        products = products.filter(category=category)

    # Attach real current stock
    for product in products:
        stock = Stock.objects.filter(product=product, store=product.store).first()
        product.current_stock = stock.quantity if stock else product.initial_stock

    return render(request, 'products/product_list.html', {
        'products': products
    })

# ===================== PRODUCT VIEWS =====================

@login_required
def Add_product(request):
    if request.method == "POST":
        # Pass user to the form
        form = ProductForm(request.POST, request.FILES, user=request.user)

        if form.is_valid():
            product = form.save(commit=False)

            # Extra safety: Force store for normal users
            if not request.user.can_see_all and request.user.store:
                product.store = request.user.store

            product.save()

            # === NOTIFICATION ===
            try:
                notify_product_added(product, created_by=request.user)
            except Exception:
                pass


            return JsonResponse({
                "success": True,
                "message": "Product added successfully!",
                "id": product.id,
                "name": product.product_name,
            })
        else:
            return JsonResponse({
                "success": False,
                "errors": dict(form.errors.items())
            }, status=400)

    # GET request
    form = ProductForm(user=request.user)   # ← Important: Pass user
    return render(request, 'products/add_product.html', {'form': form})


@login_required
def product_detail(request, id):
    product = get_object_or_404(Product.objects.for_user(request.user), id=id)
    return render(request, 'products/product_detail.html', {'product': product})


@login_required
def update_product(request, pk):
    """Update product with proper store isolation"""
    # Only allow editing products from user's store
    product = get_object_or_404(Product.objects.for_user(request.user), pk=pk)

    if request.method == "POST":
        form = ProductForm(request.POST, request.FILES, instance=product, user=request.user)

        if form.is_valid():
            updated_product = form.save(commit=False)

            # === CRITICAL FIX: Ensure store is assigned ===
            if not request.user.can_see_all and request.user.store:
                updated_product.store = request.user.store

            updated_product.save()

            # === NOTIFICATION ===
            try:
                notify_product_updated(updated_product, updated_by=request.user)
            except Exception:
                pass

            return JsonResponse({
                "success": True,
                "message": "Product updated successfully!"
            })
        else:
            print("FORM ERRORS:", form.errors)          # ← add this
            return JsonResponse({
                "success": False,
                "errors": dict(form.errors.items())
            }, status=400)

    else:
        # GET request
        form = ProductForm(instance=product, user=request.user)

    return render(request, 'products/update_product.html', {
        'form': form,
        'product': product
    })


@login_required
def delete_product(request, pk):
    product = get_object_or_404(Product.objects.for_user(request.user), pk=pk)
    product.is_deleted = True
    product.deleted_at = timezone.now()
    product.save(skip_validation=True)

    # === NOTIFICATION ===
    try:
        notify_product_deleted(
            product_name=product.product_name,
            store=product.store,
            deleted_by=request.user,
            permanent=False,
        )
    except Exception:
        pass

    messages.success(request, "Product moved to trash.")
    return redirect("product_list")

# ===================== TRASH VIEWS =====================
@login_required
def trash_products(request):
    """Show only deleted products from user's store"""
    if request.user.can_see_all:
        products = Product.all_objects.filter(is_deleted=True).order_by('-deleted_at')
    else:
        products = Product.all_objects.filter(
            store=request.user.store,
            is_deleted=True
        ).order_by('-deleted_at')

    return render(request, "products/trash.html", {"products": products})


@login_required
def restore_product(request, pk):
    """Restore product from trash"""
    if request.user.can_see_all:
        product = get_object_or_404(Product.all_objects, pk=pk)
    else:
        product = get_object_or_404(
            Product.all_objects.filter(store=request.user.store),
            pk=pk
        )

    product.is_deleted = False
    product.deleted_at = None
    product.save(skip_validation=True)
    messages.success(request, "Product restored successfully.")
    return redirect("trash_products")


@login_required
def permanent_delete(request, pk):
    """Permanently delete product + related stock history"""
    if request.user.can_see_all:
        product = get_object_or_404(Product.all_objects, pk=pk)
    else:
        product = get_object_or_404(
            Product.all_objects.filter(store=request.user.store),
            pk=pk
        )

    product_name = product.product_name
    store = product.store                      # ← FIXED: capture store before delete

    try:
        with transaction.atomic():
            # 1. Delete related Stock records
            Stock.objects.filter(product=product).delete()

            # 2. Delete related StockAdjustment history
            StockAdjustment.objects.filter(product=product).delete()

            # 3. Now delete the product itself
            product.delete()

        # === NOTIFICATION (FIXED: added) ===
        try:
            notify_product_deleted(
                product_name=product_name,
                store=store,
                deleted_by=request.user,
                permanent=True,
            )
        except Exception:
            pass

        messages.success(request, f"Product '{product_name}' and its history permanently deleted.")
    except ProtectedError as e:
        messages.error(
            request,
            f"Cannot delete '{product_name}' because it is still referenced by other records."
        )
    except Exception as e:
        messages.error(request, f"Error deleting product: {str(e)}")

    return redirect("trash_products")


@login_required
def bulk_trash_action(request):
    """Bulk restore or permanently delete from trash"""
    if request.method != "POST":
        return redirect("trash_products")

    ids = request.POST.getlist("selected_products")
    action = request.POST.get("action")

    if not ids:
        messages.error(request, "No products selected.")
        return redirect("trash_products")

    # Get the correct queryset with store isolation
    if request.user.can_see_all:
        products = Product.all_objects.filter(id__in=ids)
    else:
        products = Product.all_objects.filter(
            id__in=ids,
            store=request.user.store
        )

    if not products.exists():
        messages.error(request, "No valid products found.")
        return redirect("trash_products")

    if action == "restore":
        products.update(is_deleted=False, deleted_at=None)
        messages.success(request, f"{products.count()} product(s) restored successfully.")

    elif action == "delete":
        deleted_count = 0
        try:
            with transaction.atomic():
                for product in products:
                    product_name = product.product_name      # ← FIXED: capture
                    store = product.store                     # ← FIXED: capture

                    # Delete related stock records
                    Stock.objects.filter(product=product).delete()

                    # Delete related stock adjustment history
                    StockAdjustment.objects.filter(product=product).delete()

                    # Now delete the product
                    product.delete()
                    deleted_count += 1

                    # === NOTIFICATION (FIXED: added) ===
                    try:
                        notify_product_deleted(
                            product_name=product_name,
                            store=store,
                            deleted_by=request.user,
                            permanent=True,
                        )
                    except Exception:
                        pass

            messages.success(
                request,
                f"{deleted_count} product(s) and their history permanently deleted."
            )
        except Exception as e:
            messages.error(request, f"Error during permanent delete: {str(e)}")

    else:
        messages.error(request, "Invalid action.")

    return redirect("trash_products")

# ===================== STORE VIEWS =====================
@login_required
def store_list(request):
    """Only show stores user has access to"""
    if request.user.can_see_all:
        stores = Store.objects.all().order_by('name')
    else:
        stores = Store.objects.filter(id=request.user.store.id) if request.user.store else Store.objects.none()

    return render(request, 'store/store_list.html', {'stores': stores})


@login_required
def add_store(request):
    if not request.user.can_see_all:
        messages.error(request, "You don't have permission to add new stores. Only Administrators can do this.")
        return redirect('store_list')

    if request.method == 'POST':
        form = StoreForm(request.POST)
        if form.is_valid():
            store = form.save()                            # ← FIXED: capture store

            try:
                notify_store_created(store, created_by=request.user)
            except Exception:
                pass

            messages.success(request, "Store created successfully!")
            return redirect('store_list')
        else:
            messages.error(request, "Please correct the errors below.")
    else:
        form = StoreForm()

    return render(request, 'store/add_store.html', {'form': form})


@login_required
def edit_store(request, pk):
    """Only Admin can edit store"""
    if not request.user.can_see_all:
        messages.error(request, "Only Administrator can edit stores.")
        return redirect('store_list')

    store = get_object_or_404(Store, pk=pk)

    if request.method == 'POST':
        form = StoreForm(request.POST, instance=store)
        if form.is_valid():
            form.save()
            messages.success(request, "Store updated successfully!")
            return redirect('store_list')
    else:
        form = StoreForm(instance=store)

    return render(request, 'store/add_store.html', {
        'form': form,
        'edit_mode': True,
        'store': store
    })


@login_required
def delete_store(request, pk):
    """Only Admin can delete store + Protection"""

    if not request.user.can_see_all:
        messages.error(request, "Only Administrator can delete stores.")
        return redirect('store_list')

    store = get_object_or_404(Store, pk=pk)

    # Protection 1
    if CustomUser.objects.filter(store=store).exists():
        messages.error(request, "Cannot delete this store. Some users are currently assigned to it.")
        return redirect('store_list')

    # Protection 2
    if store.products.exists():
        messages.error(request, "Cannot delete this store. It contains products.")
        return redirect('store_list')

    # Delete
    store_name = store.name
    store.delete()
    messages.success(request, f"Store '{store_name}' deleted successfully!")

    return redirect('store_list')


# ======================================================
# Helper: Calculate Sell & Wholesale Price
# ======================================================
def calculate_prices(buy_price, category):
    if buy_price is None or buy_price <= 0:
        return Decimal('0'), Decimal('0')

    buy = Decimal(str(buy_price))
    category_name = (category or "").lower().strip()

    # Accessories
    if "accessor" in category_name:
        sell = buy + (buy * Decimal('0.35'))
        wholesale = buy + (buy * Decimal('0.20'))

    # Laptop / Desktop
    elif "laptop" in category_name or "desktop" in category_name:
        if buy <= 399999:
            sell = buy + Decimal('50000')
        elif buy <= 499999:
            sell = buy + Decimal('70000')
        elif buy <= 699999:
            sell = buy + Decimal('90000')
        elif buy <= 999999:
            sell = buy + Decimal('100000')
        elif buy <= 1499999:
            sell = buy + Decimal('150000')
        elif buy <= 1999999:
            sell = buy + Decimal('200000')
        elif buy <= 2499999:
            sell = buy + Decimal('250000')
        elif buy <= 2999999:
            sell = buy + Decimal('300000')
        else:
            sell = buy + (buy * Decimal('0.25'))

        margin = sell - buy
        wholesale = buy + (margin * Decimal('0.65'))

    # Other categories
    else:
        sell = buy + (buy * Decimal('0.25'))
        wholesale = buy + (buy * Decimal('0.15'))

    sell = sell.quantize(Decimal('1'))
    wholesale = wholesale.quantize(Decimal('1'))
    return sell, wholesale


# ======================================================
# Stock Adjustment Page
# ======================================================
@login_required
def stock_adjustment(request):
    is_admin = getattr(request.user, 'can_see_all', False) or request.user.is_superuser

    if is_admin:
        stores_qs = Store.objects.filter(is_active=True).order_by('name')
        products_qs = Product.objects.all().select_related('store')
    else:
        if request.user.store:
            stores_qs = Store.objects.filter(id=request.user.store.id)
            products_qs = Product.objects.filter(store=request.user.store)
        else:
            stores_qs = Store.objects.none()
            products_qs = Product.objects.none()

    stores = list(stores_qs.values('id', 'name'))

    products = []
    for p in products_qs:
        if p.product_group and p.variant:
            display_name = f"{p.product_group} ({p.variant})"
        else:
            display_name = p.product_name

        try:
            buy_price = float(p.buy_price) if p.buy_price is not None else 0.0
        except (TypeError, ValueError):
            buy_price = 0.0

        try:
            sell_price = float(p.sell_price) if p.sell_price is not None else 0.0
        except (TypeError, ValueError):
            sell_price = 0.0

        products.append({
            'id': p.id,
            'product_name': p.product_name,
            'display_name': display_name,
            'product_group': p.product_group or '',
            'variant': p.variant or '',
            'unit_type': p.unit_type or '',
            'buy_price': buy_price,
            'sell_price': sell_price,
            'store_id': p.store_id,
            'category': p.category or '',
        })

    return render(request, 'stock/stock_adjustment.html', {
        'stores': stores,
        'products': products,
    })


# ======================================================
# Save Stock Adjustments
# ======================================================
@login_required
@require_POST
def save_stock_adjustments(request):
    try:
        data = json.loads(request.body)
    except json.JSONDecodeError:
        return JsonResponse({'success': False, 'error': 'Invalid JSON data'}, status=400)

    if not data:
        return JsonResponse({'success': False, 'error': 'No adjustments provided'}, status=400)

    try:
        with transaction.atomic():
            for item in data:
                product = Product.objects.for_user(request.user).get(id=item['product_id'])
                store = Store.objects.get(id=item['store_id'])

                if not request.user.can_see_all:
                    if store.id != request.user.store_id or product.store_id != request.user.store_id:
                        return JsonResponse(
                            {'success': False, 'error': 'You cannot adjust this store/product.'},
                            status=403,
                        )

                quantity = int(item['quantity'])
                if quantity <= 0:
                    raise ValueError('Quantity must be a positive number.')

                adj_type = item['type']
                new_unit_price = Decimal(str(item.get('price', 0) or 0))

                stock, created = Stock.objects.get_or_create(
                    product=product, store=store, defaults={'quantity': 0}
                )

                if adj_type == 'increase':
                    old_qty = Decimal(stock.quantity)   # stock on hand BEFORE this addition
                    added_qty = Decimal(quantity)
                    total_qty = old_qty + added_qty     # never zero (added_qty >= 1)

                    stock.quantity += quantity
                    stock.save()

                    if new_unit_price > 0 and new_unit_price != product.buy_price:
                        present_buy = product.buy_price or Decimal('0')

                        # Step 1: moving-average cost
                        final_buy = (
                            (present_buy * old_qty + new_unit_price * added_qty) / total_qty
                        ).quantize(Decimal('0.01'))

                        # Step 2: derive sell & wholesale from the FINAL average buy price
                        final_sell, final_wholesale = calculate_prices(final_buy, product.category)

                        Product.objects.filter(id=product.id).update(
                            buy_price=final_buy,
                            sell_price=final_sell,
                            wholesale_price=final_wholesale,
                        )

                elif adj_type == 'decrease':
                    if stock.quantity < quantity:
                        raise ValueError(
                            f"Not enough stock for {product.product_name}. Available: {stock.quantity}"
                        )
                    stock.quantity -= quantity
                    stock.save()
                else:
                    raise ValueError(f"Invalid adjustment type: {adj_type}")

                StockAdjustment.objects.create(
                    product=product,
                    to_store=store if adj_type == 'increase' else None,
                    from_store=store if adj_type == 'decrease' else None,
                    adjustment_type=adj_type,
                    quantity=quantity,
                    unit_price=new_unit_price,
                    adjusted_by=request.user,
                )

                # === NOTIFICATION: stock adjustment ===
                try:
                    notify_stock_adjustment(
                        product=product,
                        store=store,
                        adj_type=adj_type,
                        quantity=quantity,
                        adjusted_by=request.user,
                        unit_price=new_unit_price if new_unit_price else None,
                    )
                except Exception:
                    pass

                # === NOTIFICATION: low/out of stock ===
                try:
                    check_stock_and_notify(product, store)
                except Exception:
                    pass

        return JsonResponse({'success': True, 'message': f'{len(data)} adjustment(s) saved successfully'})

    except Product.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Product not found'}, status=404)
    except Store.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Store not found'}, status=404)
    except ValueError as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=400)
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=400)

@login_required
def stock_levels(request):
    """
    Show current stock levels for products.
    Admin sees all stores, normal users see only their store.
    """
    user = request.user
    is_admin = getattr(user, 'can_see_all', False) or user.is_superuser

    # Base queryset
    if is_admin:
        stocks = Stock.objects.select_related('product', 'store', 'product__store')
        products_without_stock = Product.objects.filter(is_deleted=False)
        stocks_qs = Stock.objects.select_related('product')
    else:
        if not user.store:
            return render(request, 'stock/stock_levels.html', {
                'error': 'You are not assigned to any store.'
            })
        stocks = Stock.objects.filter(store=user.store).select_related('product', 'store')
        products_without_stock = Product.objects.filter(
            store=user.store,
            is_deleted=False
        )
        stocks_qs = Stock.objects.filter(store=user.store).select_related('product')

    # Search & filters
    search = request.GET.get('search', '').strip()
    store_id = request.GET.get('store')
    low_stock_only = request.GET.get('low_stock') == '1'

    if search:
        stocks = stocks.filter(
            Q(product__product_name__icontains=search) |
            Q(product__description__icontains=search)
        )
        products_without_stock = products_without_stock.filter(
            Q(product_name__icontains=search) |
            Q(description__icontains=search)
        )

    if is_admin and store_id:
        stocks = stocks.filter(store_id=store_id)
        products_without_stock = products_without_stock.filter(store_id=store_id)

    # Build a list of all items (products that have stock + products that don't)
    stock_data = []

    # Products that already have a Stock record
    for stock in stocks:
        current_qty = stock.quantity
        reorder = stock.product.reorder_level or 0
        stock_data.append({
            'product': stock.product,
            'store': stock.store,
            'quantity': current_qty,
            'reorder_level': reorder,
            'is_low': current_qty <= reorder,
            'stock_id': stock.id,
        })

    # Products that never had a Stock record yet (use initial_stock)
    existing_product_ids = stocks.values_list('product_id', flat=True)
    for product in products_without_stock.exclude(id__in=existing_product_ids):
        current_qty = product.initial_stock or 0
        reorder = product.reorder_level or 0
        stock_data.append({
            'product': product,
            'store': product.store,
            'quantity': current_qty,
            'reorder_level': reorder,
            'is_low': current_qty <= reorder,
            'stock_id': None,
        })

    # Filter low stock only
    if low_stock_only:
        stock_data = [item for item in stock_data if item['is_low']]

    # Sort by product name
    stock_data.sort(key=lambda x: x['product'].product_name.lower())

    # Pagination
    paginator = Paginator(stock_data, 25)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    # For admin filter dropdown
    stores = Store.objects.filter(is_active=True).order_by('name') if is_admin else []

    # Summary cards
    total_items = len(stock_data)
    low_stock_count = sum(1 for item in stock_data if item['is_low'])
    total_quantity = sum(item['quantity'] for item in stock_data)

    stock_value = stocks_qs.aggregate(
            total=Sum(F('quantity') * F('product__buy_price'))
        )['total'] or Decimal('0')

    context = {
        'page_obj': page_obj,
        'stock_value': stock_value,
        'stock_data': page_obj,          # for easier template access
        'is_admin': is_admin,
        'stores': stores,
        'search': search,
        'selected_store': store_id,
        'low_stock_only': low_stock_only,
        'total_items': total_items,
        'low_stock_count': low_stock_count,
        'total_quantity': total_quantity,
        'current_store': user.store if not is_admin else None,
    }
    return render(request, 'stock/stock_levels.html', context)




def _get_stock_export_data(request):
    """
    Shared logic to get the same stock data as stock_levels view.
    Respects admin vs normal user and all filters.
    """
    user = request.user
    is_admin = getattr(user, 'can_see_all', False) or user.is_superuser

    if is_admin:
        stocks = Stock.objects.select_related('product', 'store', 'product__store')
        products_without_stock = Product.objects.filter(is_deleted=False)
    else:
        if not user.store:
            return []
        stocks = Stock.objects.filter(store=user.store).select_related('product', 'store')
        products_without_stock = Product.objects.filter(
            store=user.store,
            is_deleted=False
        )

    # Filters (same as stock_levels)
    search = request.GET.get('search', '').strip()
    store_id = request.GET.get('store')
    low_stock_only = request.GET.get('low_stock') == '1'

    if search:
        stocks = stocks.filter(
            Q(product__product_name__icontains=search) |
            Q(product__description__icontains=search) |
            Q(product__product_group__icontains=search) |
            Q(product__variant__icontains=search)
        )
        products_without_stock = products_without_stock.filter(
            Q(product_name__icontains=search) |
            Q(description__icontains=search) |
            Q(product_group__icontains=search) |
            Q(variant__icontains=search)
        )

    if is_admin and store_id:
        stocks = stocks.filter(store_id=store_id)
        products_without_stock = products_without_stock.filter(store_id=store_id)

    stock_data = []

    # Products that already have Stock records
    for stock in stocks:
        product = stock.product
        current_qty = stock.quantity
        reorder = product.reorder_level or 0
        stock_data.append({
            'product_name': product.display_name,
            'store': stock.store.name,
            'quantity': current_qty,
            'buy_price': product.buy_price or 0,
            'sell_price': product.sell_price or 0,
            'is_low': current_qty <= reorder,
            'unit_type': product.get_unit_type_display() if product.unit_type else product.unit_type or '',
            'category': product.category or '',
        })

    # Products without Stock record yet
    existing_product_ids = stocks.values_list('product_id', flat=True)
    for product in products_without_stock.exclude(id__in=existing_product_ids):
        current_qty = product.initial_stock or 0
        reorder = product.reorder_level or 0
        stock_data.append({
            'product_name': product.display_name,
            'store': product.store.name,
            'quantity': current_qty,
            'buy_price': product.buy_price or 0,
            'sell_price': product.sell_price or 0,
            'is_low': current_qty <= reorder,
            'unit_type': product.get_unit_type_display() if product.unit_type else product.unit_type or '',
            'category': product.category or '',
        })

    if low_stock_only:
        stock_data = [item for item in stock_data if item['is_low']]

    # Sort
    stock_data.sort(key=lambda x: x['product_name'].lower())
    return stock_data, is_admin


@login_required
def export_stock_excel(request):
    stock_data, is_admin = _get_stock_export_data(request)

    wb = Workbook()
    ws = wb.active
    ws.title = "Stock Levels"

    # Styles
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="647688", end_color="647688", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    center_align = Alignment(horizontal='center', vertical='center')
    left_align = Alignment(horizontal='left', vertical='center')
    money_align = Alignment(horizontal='right', vertical='center')

    # Title
    ws.merge_cells('A1:G1')
    ws['A1'] = "STOCK LEVELS REPORT"
    ws['A1'].font = Font(bold=True, size=16, color="647688")
    ws['A1'].alignment = Alignment(horizontal='center')

    # Subtitle
    ws.merge_cells('A2:G2')
    generated = datetime.now().strftime("%d %b %Y %H:%M")
    store_info = "All Stores" if is_admin else (request.user.store.name if request.user.store else "N/A")
    ws['A2'] = f"Generated: {generated}  |  Store: {store_info}"
    ws['A2'].alignment = Alignment(horizontal='center')
    ws['A2'].font = Font(size=10, italic=True)

    # Headers
    headers = ["#", "Product", "Store", "Quantity", "Unit", "Buy Price (Tsh)", "Sell Price (Tsh)"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border

    # Data rows
    for idx, item in enumerate(stock_data, 1):
        row = idx + 4
        values = [
            idx,
            item['product_name'],
            item['store'],
            item['quantity'],
            item['unit_type'],
            float(item['buy_price']),
            float(item['sell_price']),
        ]
        for col, value in enumerate(values, 1):
            cell = ws.cell(row=row, column=col, value=value)
            cell.border = thin_border
            if col in (1, 4):
                cell.alignment = center_align
            elif col in (6, 7):
                cell.alignment = money_align
                cell.number_format = '#,##0.00'
            else:
                cell.alignment = left_align

            # Highlight low stock
            if item['is_low']:
                cell.fill = PatternFill(start_color="FFCDD2", end_color="FFCDD2", fill_type="solid")

    # Column widths
    column_widths = [6, 40, 18, 12, 12, 18, 18]
    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width

    # Freeze header
    ws.freeze_panes = 'A5'

    # Response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f"stock_levels_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response


@login_required
def export_stock_pdf(request):
    stock_data, is_admin = _get_stock_export_data(request)

    buffer = BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=landscape(A4),
        rightMargin=15*mm,
        leftMargin=15*mm,
        topMargin=15*mm,
        bottomMargin=15*mm
    )

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=16,
        textColor=colors.HexColor('#647688'),
        alignment=TA_CENTER,
        spaceAfter=6
    )
    subtitle_style = ParagraphStyle(
        'Subtitle',
        parent=styles['Normal'],
        fontSize=9,
        alignment=TA_CENTER,
        textColor=colors.grey,
        spaceAfter=12
    )

    elements = []

    # Title
    elements.append(Paragraph("STOCK LEVELS REPORT", title_style))

    generated = datetime.now().strftime("%d %b %Y %H:%M")
    store_info = "All Stores" if is_admin else (request.user.store.name if request.user.store else "N/A")
    elements.append(Paragraph(f"Generated: {generated}  |  Store: {store_info}", subtitle_style))
    elements.append(Spacer(1, 8))

    # Table data
    header = ["#", "Product", "Store", "Qty", "Unit", "Buy Price", "Sell Price"]
    data = [header]

    for idx, item in enumerate(stock_data, 1):
        data.append([
            str(idx),
            item['product_name'][:45],
            item['store'][:18],
            str(item['quantity']),
            item['unit_type'][:10],
            f"{float(item['buy_price']):,.0f}",
            f"{float(item['sell_price']):,.0f}",
        ])

    if not stock_data:
        data.append(["", "No stock records found.", "", "", "", "", ""])

    # Column widths (landscape A4 ≈ 270mm usable)
    col_widths = [12*mm, 75*mm, 35*mm, 18*mm, 22*mm, 30*mm, 30*mm]

    table = Table(data, colWidths=col_widths, repeatRows=1)

    style_commands = [
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#647688')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 9),
        ('FONTSIZE', (0, 1), (-1, -1), 8),
        ('ALIGN', (0, 0), (0, -1), 'CENTER'),
        ('ALIGN', (3, 0), (3, -1), 'CENTER'),
        ('ALIGN', (5, 0), (-1, -1), 'RIGHT'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('GRID', (0, 0), (-1, -1), 0.4, colors.HexColor('#CCCCCC')),
        ('TOPPADDING', (0, 0), (-1, -1), 4),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
        ('LEFTPADDING', (0, 0), (-1, -1), 4),
        ('RIGHTPADDING', (0, 0), (-1, -1), 4),
    ]

    # Highlight low stock rows
    for i, item in enumerate(stock_data, 1):
        if item['is_low']:
            style_commands.append(
                ('BACKGROUND', (0, i), (-1, i), colors.HexColor('#FFCDD2'))
            )

    table.setStyle(TableStyle(style_commands))
    elements.append(table)

    # Footer note
    elements.append(Spacer(1, 10))
    note_style = ParagraphStyle(
        'Note',
        parent=styles['Normal'],
        fontSize=8,
        textColor=colors.grey,
        alignment=TA_LEFT
    )
    elements.append(Paragraph(
        f"Total products: {len(stock_data)}  •  Pink rows = Low Stock",
        note_style
    ))

    doc.build(elements)
    buffer.seek(0)

    response = HttpResponse(buffer, content_type='application/pdf')
    filename = f"stock_levels_{datetime.now().strftime('%Y%m%d_%H%M')}.pdf"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response

def live_stock_records(request):
    return render(request, 'stock/livestock.html')